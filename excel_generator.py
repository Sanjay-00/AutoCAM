"""
excel_generator.py
Generates a formatted Excel file from extracted CIBIL data.
Matches the Shriram Finance analyst format exactly.
"""

import io
import re
import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# COLOUR PALETTE  (Shriram Finance theme)
# ─────────────────────────────────────────────

NAVY          = "1F3864"
WHITE         = "FFFFFF"
LIGHT_GREY    = "F2F2F2"
ALT_WHITE     = "FFFFFF"
SCORE_GREEN   = "C6EFCE"   # score > 700
SCORE_ORANGE  = "FFEB9C"   # score 600-700
SCORE_RED     = "FFC7CE"   # score < 600
ACTIVE_GREEN  = "375623"   # Active text colour
CLOSED_GREY   = "595959"   # Closed text colour
WRITEOFF_RED  = "C00000"   # Written Off text colour
SETTLED_AMBER = "9C6500"   # Settled text colour
DPD_CLEAR     = "C6EFCE"   # max DPD = 0

STATUS_COLORS = {
    "Active":      ACTIVE_GREEN,
    "Closed":      CLOSED_GREY,
    "Written Off": WRITEOFF_RED,
    "Settled":     SETTLED_AMBER,
}

# DPD gradient: 4 bands that intensify orange → red as DPD rises
_DPD_BANDS = [
    (30,  None,     None),        # 1-30   : no colour
    (60,  "FCE4D6", "843C0C"),    # 31-60  : light orange
    (90,  "F4B183", "843C0C"),    # 61-90  : medium orange
    (180, "FF7F7F", "7B0000"),    # 91-180 : light red
    (None,"C00000", "FFFFFF"),    # 181+   : deep red, white text
]


def _dpd_style(dpd: int):
    """Return (bg_hex, font_hex) for a DPD value, or (None, None) for 1-30."""
    if dpd == 0:
        return DPD_CLEAR, "375623"
    for cap, bg, fg in _DPD_BANDS:
        if cap is None or dpd <= cap:
            return bg, fg
    return "C00000", "FFFFFF"
TOTAL_BG      = "FFF2CC"   # total row
KP_BG         = "EBF3FB"   # key points rows
KP_HDR_BG     = "D6E4F0"   # key points section header
BORDER_CLR    = "BFBFBF"


# ─────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────

def _b(style="thin"):
    s = Side(style=style, color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def _f(size=10, bold=False, color="000000", italic=False, underline=None):
    return Font(name="Arial", size=size, bold=bold,
                color=color, italic=italic, underline=underline)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _a(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def _fmt_inr(value) -> str:
    """Format a number in Indian comma notation (e.g. 11,85,657)."""
    try:
        v = int(value)
        if v == 0:
            return "0"
        # Indian format: last 3 digits, then groups of 2
        s = str(abs(v))
        if len(s) <= 3:
            result = s
        else:
            result = s[-3:]
            s = s[:-3]
            while s:
                result = s[-2:] + "," + result
                s = s[:-2]
        return ("-" if v < 0 else "") + result
    except (TypeError, ValueError):
        return str(value)


# ─────────────────────────────────────────────
# COLUMN DEFINITIONS
# ─────────────────────────────────────────────

COLUMNS = [
    ("Sr.No",              8),
    ("Date of Sanction",  18),
    ("Sanction Amount (₹)", 20),
    ("Current Balance (₹)", 20),
    ("EMI (₹)",            14),
    ("Overdue (₹)",        14),
    ("Entity",             20),
    ("Ownership",          14),
    ("Type of Loan",       30),
    ("Max DPD",            10),
    ("Status",             24),
]


# ─────────────────────────────────────────────
# MAIN GENERATOR
# ─────────────────────────────────────────────

def generate_excel(data: dict) -> bytes:
    """
    Generate a formatted Excel file from extracted CIBIL data.
    Returns bytes that can be written to file or sent via Streamlit download.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "CIBIL Summary"

    accounts   = data.get("accounts", [])
    name       = data.get("name", data.get("borrower_name", "Unknown"))
    score      = data.get("score", data.get("cibil_score",  "NA"))
    key_points = data.get("key_points", [])

    # ── Set column widths
    for col_idx, (_, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ────────────────────────────────────────
    # ROW 1  -  Borrower Name
    # ────────────────────────────────────────
    ws["A1"] = "Borrower Name:"
    ws["A1"].font      = _f(size=12, bold=True, color=NAVY)
    ws["A1"].alignment = _a()

    ws["B1"] = name
    ws["B1"].font      = _f(size=14, bold=True, color=NAVY)
    ws["B1"].alignment = _a()
    ws.row_dimensions[1].height = 24

    # ────────────────────────────────────────
    # ROW 2  -  CIBIL Score
    # ────────────────────────────────────────
    ws["A2"] = "CIBIL Score:"
    ws["A2"].font      = _f(size=12, bold=True, color=NAVY)
    ws["A2"].alignment = _a()

    # CRIF Commercial isn't a 300-900 score - it's a 1 (best) - 5 (worst) risk
    # rank rendered as "N (Label)", e.g. "5 (Very High Risk)". Colour it on
    # its own scale (low rank = good = green) instead of falling through to
    # the flat amber "unparseable" branch below, which used to colour every
    # commercial risk rank identically regardless of how bad it actually is.
    rank_m = re.match(r'^\s*(\d)\s*\(', str(score))
    if rank_m:
        rank          = int(rank_m.group(1))
        score_display = score
        if rank <= 2:
            score_fill, score_color = _fill(SCORE_GREEN),  "375623"
        elif rank == 3:
            score_fill, score_color = _fill(SCORE_ORANGE), "7F6000"
        else:
            score_fill, score_color = _fill(SCORE_RED),    "9C0006"
    else:
        try:
            score_val = int(score)
            score_display = score_val
            if score_val > 700:
                score_fill  = _fill(SCORE_GREEN)
                score_color = "375623"
            elif score_val >= 600:
                score_fill  = _fill(SCORE_ORANGE)
                score_color = "7F6000"
            else:
                score_fill  = _fill(SCORE_RED)
                score_color = "9C0006"
        except (TypeError, ValueError):
            score_display = str(score)
            score_fill    = _fill(SCORE_ORANGE)
            score_color   = "7F6000"

    ws["B2"] = score_display
    ws["B2"].font      = _f(size=14, bold=True, color=score_color)
    ws["B2"].fill      = score_fill
    ws["B2"].alignment = _a("left")
    ws.row_dimensions[2].height = 22

    # ────────────────────────────────────────
    # ROW 3  -  Empty spacer
    # ────────────────────────────────────────
    ws.row_dimensions[3].height = 8

    # ────────────────────────────────────────
    # ROW 4  -  Table Headers
    # ────────────────────────────────────────
    hdr_fill = _fill(NAVY)
    for col_idx, (header, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font      = _f(size=11, bold=True, color=WHITE)
        cell.fill      = hdr_fill
        cell.alignment = _a("center", wrap=True)
        cell.border    = _b()
    ws.row_dimensions[4].height = 30

    # ────────────────────────────────────────
    # ROWS 5+  -  Account Data
    # ────────────────────────────────────────
    DATA_START = 5
    for idx, acc in enumerate(accounts):
        row_num = DATA_START + idx

        # Alternating row background
        row_bg = _fill(LIGHT_GREY) if idx % 2 == 1 else _fill(ALT_WHITE)

        # DPD value used for gradient colouring when it's a confident read
        dpd_raw = acc.get("max_dpd")
        try:
            dpd = int(dpd_raw)
        except (TypeError, ValueError):
            dpd = 0

        # Prepare cell values
        date_val = acc.get("date_of_sanction", "NA")
        # Try to parse date for proper Excel date type
        date_cell_val = _parse_date(date_val)

        # None signals OCR read failure — shown as "Check CIBIL" in output
        _CHECK = "Check CIBIL"
        sanction_val = acc.get("sanction_amount")
        balance_val  = acc.get("current_balance")
        sanction_cell = _CHECK if sanction_val is None else sanction_val
        balance_cell  = _CHECK if balance_val  is None else balance_val
        dpd_cell      = _CHECK if dpd_raw     is None else dpd

        # CRIF Commercial only: "Active"/"Closed"/"Written Off"/"Settled" is the
        # canonical status (used for validation math elsewhere); Delinquent and
        # Suit Filed are independent overlay flags, appended for the analyst's
        # benefit without disturbing that canonical value.
        status_raw = acc.get("status", "Active")
        flags = []
        if acc.get("delinquent"):
            flags.append("Delinquent")
        if acc.get("suit_filed"):
            flags.append("Suit Filed")
        status_cell = f"{status_raw} ({', '.join(flags)})" if flags else status_raw

        row_values = [
            acc.get("sr_no", idx + 1),
            date_cell_val,
            sanction_cell,
            balance_cell,
            acc.get("emi", 0),
            acc.get("overdue", 0),
            acc.get("entity", "XXXX"),
            acc.get("ownership", ""),
            acc.get("type_of_loan", ""),
            dpd_cell,
            status_cell,
        ]

        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border    = _b()

            # Column-specific formatting
            if col_idx == 1:   # Sr.No
                cell.fill      = row_bg
                cell.font      = _f(bold=True)
                cell.alignment = _a("center")

            elif col_idx == 2:   # Date
                cell.fill      = row_bg
                cell.font      = _f()
                cell.alignment = _a("center")
                if isinstance(value, datetime.datetime):
                    cell.number_format = "DD-MMM-YYYY"

            elif col_idx in (3, 4, 5, 6):   # Amount columns
                if value == _CHECK:
                    cell.fill      = _fill("FFF2CC")   # amber tint
                    cell.font      = _f(italic=True, color="7F6000", bold=True)
                    cell.alignment = _a("center")
                else:
                    cell.fill          = row_bg
                    cell.font          = _f()
                    cell.alignment     = _a("right")
                    cell.number_format = '#,##0;-#,##0;"-"'

            elif col_idx == 7:   # Entity
                cell.fill      = row_bg
                cell.font      = _f()
                cell.alignment = _a("left")

            elif col_idx == 8:   # Ownership
                cell.fill      = row_bg
                cell.font      = _f()
                cell.alignment = _a("center")

            elif col_idx == 9:   # Type of Loan
                cell.fill      = row_bg
                cell.font      = _f()
                cell.alignment = _a("left", wrap=True)

            elif col_idx == 10:   # Max DPD - gradient colour
                if value == _CHECK:
                    cell.fill      = _fill("FFF2CC")   # amber tint
                    cell.font      = _f(italic=True, color="7F6000", bold=True)
                    cell.alignment = _a("center")
                else:
                    bg, fg = _dpd_style(dpd)
                    if bg is None:          # 1-30: plain
                        cell.fill = row_bg
                        cell.font = _f()
                    else:
                        cell.fill = _fill(bg)
                        cell.font = _f(color=fg, bold=True)
                    cell.alignment = _a("center")

            elif col_idx == 11:   # Status
                status_color = STATUS_COLORS.get(status_raw, CLOSED_GREY)
                if acc.get("delinquent"):
                    # Amber tint (same as 'Check CIBIL' cells) so a delinquent
                    # account stands out at a glance, not just in the "(Delinquent)"
                    # text - matches the app's row highlight for the same accounts.
                    cell.fill = _fill("FFF2CC")
                    cell.font = _f(color="7F6000", bold=True)
                else:
                    cell.fill = row_bg
                    cell.font = _f(color=status_color, bold=(status_raw == "Active"))
                cell.alignment = _a("center")

        ws.row_dimensions[row_num].height = 16

    # ────────────────────────────────────────
    # TOTAL ROW (Current Balance sum)
    # ────────────────────────────────────────
    last_data    = DATA_START + len(accounts) - 1
    total_row    = DATA_START + len(accounts)

    label_cell = ws.cell(row=total_row, column=3, value="TOTAL ACTIVE EXPOSURE →")
    label_cell.font      = _f(bold=True, color=NAVY)
    label_cell.alignment = _a("right")
    label_cell.fill      = _fill(TOTAL_BG)
    label_cell.border    = _b()

    sum_cell = ws.cell(row=total_row, column=4,
                       value=f"=SUMIF(K{DATA_START}:K{last_data},\"Active\",D{DATA_START}:D{last_data})")
    sum_cell.number_format = '#,##0'
    sum_cell.font          = _f(bold=True, color=NAVY)
    sum_cell.alignment     = _a("right")
    sum_cell.fill          = _fill(TOTAL_BG)
    sum_cell.border        = _b()
    ws.row_dimensions[total_row].height = 18

    # ────────────────────────────────────────
    # KEY POINTS SECTION
    # ────────────────────────────────────────
    kp_gap_start  = total_row + 1
    ws.row_dimensions[kp_gap_start].height     = 8
    ws.row_dimensions[kp_gap_start + 1].height = 8

    kp_header_row = kp_gap_start + 2
    ws.merge_cells(f"A{kp_header_row}:K{kp_header_row}")
    kp_hdr = ws.cell(row=kp_header_row, column=1,
                     value="Key Points for Loan Decision:")
    kp_hdr.font      = _f(size=12, bold=True, underline="single", color=NAVY)
    kp_hdr.fill      = _fill(KP_HDR_BG)
    kp_hdr.alignment = _a(indent=1)
    ws.row_dimensions[kp_header_row].height = 22

    for i, point in enumerate(key_points):
        pr = kp_header_row + 1 + i
        ws.merge_cells(f"A{pr}:K{pr}")
        kp_cell = ws.cell(row=pr, column=1, value=f"  {i+1}. {point}")
        kp_cell.font      = _f(size=10)
        kp_cell.fill      = _fill(KP_BG)
        kp_cell.alignment = _a(wrap=True, indent=1)
        ws.row_dimensions[pr].height = 22

    # ────────────────────────────────────────
    # FREEZE PANES  (freeze rows 1-4 including header)
    # ────────────────────────────────────────
    ws.freeze_panes = "A5"

    # ────────────────────────────────────────
    # PRINT SETTINGS
    # ────────────────────────────────────────
    ws.print_title_rows   = "1:4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1

    # ────────────────────────────────────────
    # CREDIT ANALYSIS SHEET  (CRIF Commercial only - portfolio-level view
    # above the individual accounts: Shriram's own exposure vs the market,
    # asset-class distribution, red-flag rollup)
    # ────────────────────────────────────────
    analysis = data.get("analysis")
    if analysis:
        _build_credit_analysis_sheet(wb, analysis)

    # ────────────────────────────────────────
    # SAVE TO BYTES BUFFER
    # ────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ─────────────────────────────────────────────
# CREDIT ANALYSIS SHEET
# ─────────────────────────────────────────────

_ASSET_CLASS_STYLE = {
    "Standard":      (DPD_CLEAR,  "375623"),
    "SMA-0":         (None,       "000000"),
    "SMA-1":         ("FCE4D6",   "843C0C"),
    "SMA-2":         ("F4B183",   "843C0C"),
    "Substandard":   ("FF7F7F",   "7B0000"),
    "Doubtful/Loss": ("C00000",   "FFFFFF"),
    "Unclassified":  (None,       "595959"),
}


def _section_header(ws, row, text, span="A:F"):
    start_col, end_col = span.split(":")
    ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = _f(size=12, bold=True, color=WHITE)
    cell.fill      = _fill(NAVY)
    cell.alignment = _a(indent=1)
    ws.row_dimensions[row].height = 22
    return row + 1


def _kv_row(ws, row, label, value, number_format=None):
    lbl = ws.cell(row=row, column=1, value=label)
    lbl.font      = _f(bold=True, color=NAVY)
    lbl.alignment = _a()
    val = ws.cell(row=row, column=2, value=value if value is not None else "NA")
    val.font = _f()
    if number_format and isinstance(value, (int, float)):
        val.number_format = number_format
    return row + 1


def _build_credit_analysis_sheet(wb, analysis: dict) -> None:
    ws = wb.create_sheet("Credit Analysis")
    for col, width in zip("ABCDEFGHI", (28, 20, 18, 18, 20, 20, 22, 18, 18)):
        ws.column_dimensions[col].width = width

    row = 1
    title = ws.cell(row=row, column=1, value="Credit Analysis")
    title.font = _f(size=14, bold=True, color=NAVY)
    row += 2

    # ── Borrower Summary: Shriram's own book vs the rest of the market ──
    # CRIF Commercial only - Retail's report format has no market-comparison
    # table to pull this from at all (borrower_summary key is simply absent
    # for Retail's analysis dict), so this whole section is skipped rather
    # than rendered full of misleading "NA" rows.
    bs = analysis.get("borrower_summary") or {}
    if bs:
        row = _section_header(ws, row, "Borrower Summary  -  Our Exposure vs Market")

        headers = ["", "Lenders", "Total Accts", "Live Accts", "Delinquent Accts",
                  "Sanctioned (Rs.)", "Outstanding (Rs.)", "Overdue (Rs.)", "PAR 90+ (Rs.)"]
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col_idx, value=h)
            c.font = _f(bold=True, color=NAVY)
            c.fill = _fill(LIGHT_GREY)
            c.border = _b()
        row += 1

        for label, key in (("Your Institution (Shriram)", "your_institution"),
                           ("Other Institutions (Market)", "other_institution")):
            inst = bs.get(key) or {}
            values = [
                label,
                inst.get("lenders"), inst.get("total_accts"), inst.get("live_accts"),
                inst.get("delinquent_accts"),
                inst.get("sanctioned_amt"), inst.get("outstanding_amt"),
                inst.get("overdue_amt"), inst.get("par_90plus"),
            ]
            for col_idx, v in enumerate(values, 1):
                c = ws.cell(row=row, column=col_idx, value=v if v is not None else "NA")
                c.border = _b()
                if col_idx == 1:
                    c.font = _f(bold=True)
                elif col_idx >= 6:
                    c.number_format = '#,##0;-#,##0;"NA"'
                    c.alignment = _a("right")
            row += 1
        row += 1

        row = _kv_row(ws, row, "Length of Credit History:", bs.get("length_of_credit_history"))
        row = _kv_row(ws, row, "New Accounts (last 12 months):", bs.get("new_accts_12m"))
        row = _kv_row(ws, row, "Closed Accounts (last 12 months):", bs.get("closed_accts_12m"))
        row = _kv_row(ws, row, "New Delinquent Accounts (last 12 months):", bs.get("new_delinquent_accts_12m"))
        row += 2

    # ── Credit Profile Summary: asset-class (Commercial) / loan-type (Retail)
    #    distribution - same shape either way, just a different label.
    section_label = ("Credit Profile Summary  -  Asset-Class Distribution (Active Accounts)"
                     if bs else
                     "Credit Profile Summary  -  Loan-Type Distribution (Active Accounts)")
    row = _section_header(ws, row, section_label)
    col1_header = "Asset Class" if bs else "Loan Type"
    for col_idx, h in enumerate([col1_header, "Accounts", "Outstanding Balance (Rs.)"], 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = _f(bold=True, color=NAVY)
        c.fill = _fill(LIGHT_GREY)
        c.border = _b()
    row += 1

    cps = analysis.get("credit_profile_summary") or []
    for bucket in cps:
        bg, fg = _ASSET_CLASS_STYLE.get(bucket["asset_class"], (None, "000000"))
        cells = [bucket["asset_class"], bucket["count"], bucket["outstanding"]]
        for col_idx, v in enumerate(cells, 1):
            c = ws.cell(row=row, column=col_idx, value=v)
            c.border = _b()
            c.font   = _f(bold=(col_idx == 1), color=fg)
            if bg:
                c.fill = _fill(bg)
            if col_idx == 3:
                c.number_format = '#,##0'
                c.alignment = _a("right")
        row += 1
    if not cps:
        ws.cell(row=row, column=1, value="No active accounts to classify").font = _f(italic=True)
        row += 1
    row += 1

    # ── Derogatory Status Summary ──
    row = _section_header(ws, row, "Derogatory Status Summary")
    for col_idx, h in enumerate(["Category", "Accounts", "Amount (Rs.)"], 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = _f(bold=True, color=NAVY)
        c.fill = _fill(LIGHT_GREY)
        c.border = _b()
    row += 1

    derog = analysis.get("derog_summary") or {}
    # Written Off / Suit Filed show original exposure (sanction amount) - the
    # bureau zeroes current_balance once an account reaches either status, so
    # a plain "Amount" column would otherwise read as "Rs.0 impact". Settled /
    # Delinquent show current balance, which is still meaningful for those two.
    # Rows are driven by which keys the analysis dict actually has - Retail's
    # derog_summary() only ever returns written_off/overdue (no Settled/Suit
    # Filed/Delinquent concept in that report format), Commercial's returns
    # all four of its own set. Order is fixed regardless of which subset shows.
    _DEROG_ROW_STYLE = {
        "written_off": ("Written Off (orig. exposure)", WRITEOFF_RED),
        "settled":     ("Settled (balance)",             SETTLED_AMBER),
        "suit_filed":  ("Suit Filed (orig. exposure)",   WRITEOFF_RED),
        "delinquent":  ("Delinquent (balance)",          "843C0C"),
        "overdue":     ("Overdue (active accounts)",     "843C0C"),
    }
    for key in ("written_off", "settled", "suit_filed", "delinquent", "overdue"):
        if key not in derog:
            continue
        label, color = _DEROG_ROW_STYLE[key]
        d = derog[key]
        cells = [label, d["count"], d["amount"]]
        for col_idx, v in enumerate(cells, 1):
            c = ws.cell(row=row, column=col_idx, value=v)
            c.border = _b()
            c.font   = _f(bold=(col_idx == 1), color=color if col_idx == 1 else "000000")
            if col_idx == 3:
                c.number_format = '#,##0'
                c.alignment = _a("right")
        row += 1

    ws.freeze_panes = "A2"


def _parse_date(date_str):
    """
    Try to parse a date string into a datetime object.
    Falls back to the original string if parsing fails.
    """
    if not date_str or date_str in ("NA", "N/A", ""):
        return "NA"
    formats = ["%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y"]
    for fmt in formats:
        try:
            return datetime.datetime.strptime(str(date_str), fmt)
        except ValueError:
            continue
    return date_str   # return as-is if unparseable


def get_filename(borrower_name: str) -> str:
    """Generate output filename with borrower name and today's date."""
    safe_name = re.sub(r"[^\w\s-]", "", borrower_name).strip().replace(" ", "_")
    date_str  = datetime.datetime.now().strftime("%d%b%Y")
    return f"CIBIL_Analysis_{safe_name}_{date_str}.xlsx"


import re   # needed for get_filename
